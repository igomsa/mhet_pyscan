#!/usr/bin/env python3
#
from datetime import datetime
from openpyxl import Workbook
from dateutil.parser import *
import openpyxl
import re

base_datos = "source_path"
lista_asistencia = "output_path"
#workbook = Workbook()
#sheet = workbook.active


#####################################################3
def buscar_codigo(codigo):
    wb = openpyxl.load_workbook(base_datos)
    ws = wb.active
    i=0
    for cell in ws['B']:
        i += 1
        if cell.value == codigo:
                #print(i)
                #print(cell.value)
                #print(ws.cell(i,1).value)
                nombre = ws.cell(i,1).value
                return nombre
    return 0
    #nombre = "valor no encontrado"
    #return nombre

#####################################################3
def revisar_fecha(wo, wc):

    fecha_hoy = datetime.today().strftime('%d-%m-%Y')
    print("\nFecha de hoy: " + fecha_hoy)
    k=0
    for cell in wo[1]:
        if cell.value:
            k += 1
            # searching string
            #if  (cell.value):
            try:
                fecha_celda=parse(cell.value, dayfirst=True).strftime('%d-%m-%Y')
                print("\nFecha encontrada en celda: " + fecha_celda)
            except:
                continue
            if fecha_celda == fecha_hoy:
                print("\nFechas iguales")
                #print("wo max col: " + str(k))
                return k
    k += 1
    print("\nFecha diferentes")
    wo.cell(row = 1, column=k, value=fecha_hoy)
    wc.save(filename="docs/asistencia.xlsx")
    return k

#####################################################3
def asistencia(nombre):
    wc = openpyxl.load_workbook(lista_asistencia)
    wo = wc.active
    i=0
    for cell in wo['A']:
        if cell.value:
            i += 1
            #print(cell.value)
            if cell.value == nombre:
                print("\nNombre encontrado")
                j = revisar_fecha(wo,wc)
                #print("i: " + str(i))
                #print("j: " + str(j))
                wo.cell(row=i,column=j, value="x")
                wc.save(filename="docs/asistencia.xlsx")
                return 1
    print("\nNombre  no encontrado")
    j = revisar_fecha(wo, wc)
    #print("i: " + str(i+1))
    #print("j: " + str(j))
    wo.cell(row=i+1,column=1, value=nombre)
    wo.cell(row=i+1,column=j, value="x")
    wc.save(filename="docs/asistencia.xlsx")
    return 1
